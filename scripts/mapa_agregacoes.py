"""Aplica o mapa de agregacoes do TSE a Dublin e gera a planilha de analise.

Produz saidas/Dublin_2026_agregacoes.xlsx com cinco abas e saidas/dados.json
com o mesmo conteudo para alimentar a pagina visual.

O criterio de gargalo NAO e aplicado aqui: a planilha entrega os totais
ordenados e quem define o corte e o usuario.
"""

import json
from pathlib import Path

import pandas as pd
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from parse_dados import (
    carrega_locais_votacao,
    carrega_perfil_eleitorado,
    residencia_por_secao,
)

BASE = Path(__file__).resolve().parent.parent
SAIDAS = BASE / "saidas"

# Mapa transcrito do PNG "2026.8.13_Mapa de agregacoes_Final_TSE.png".
# A linha da secao 3752 traz "3222" no PNG, mas 3222 e uma secao do PORTO --
# a secao de Dublin e a 3322. Tratado como erro de digitacao do mapa; ver a
# aba "Inconsistencias" da planilha.
MAPA_PNG = {
    1100: 511, 2855: 512, 1105: 513, 1292: 517, 3845: 1160, 522: 1352,
    1099: 3054, 2847: 3078, 3422: 3108, 1278: 3142, 3307: 3161, 530: 3179,
    527: 3216, 3821: 3229, 519: 3245, 3181: 3302, 521: 3305, 518: 3306,
    1314: 3309, 3913: 3311, 3889: 3313, 3778: 3315, 3752: 3222,
}

CABECALHO = PatternFill("solid", fgColor="1F3864")
FONTE_CABECALHO = Font(bold=True, color="FFFFFF", size=11)


def monta_secoes(locais: pd.DataFrame, residencia: pd.DataFrame) -> pd.DataFrame:
    """Uma linha por secao, com papel, urna de destino e eleitorado."""
    perfil_total = residencia.groupby("NR_SECAO")["QT_ELEITORES"].sum()

    df = locais[[
        "NR_SECAO", "DS_TIPO_SECAO_AGREGADA", "NR_SECAO_PRINCIPAL",
        "QT_ELEITOR_SECAO", "QT_ELEITOR_ELEICAO_FEDERAL",
    ]].copy()
    # QT_ELEITOR_ELEICAO_FEDERAL traz, na secao principal, o total ja agregado
    # que votara naquela urna; nas secoes agregadas vem zerado. Serve de
    # conferencia independente do total_combinado calculado aqui.
    df.columns = ["Secao", "Papel", "Secao_principal_bruta",
                  "Eleitores", "Total_urna_TSE"]

    # Para uma secao principal a urna e ela mesma; para uma agregada e a principal.
    df["Urna"] = df.apply(
        lambda r: int(r["Secao"]) if r["Papel"] == "Principal"
        else int(r["Secao_principal_bruta"]),
        axis=1,
    )
    df["Eleitores_perfil"] = df["Secao"].map(perfil_total).fillna(0).astype(int)
    df["Diferenca"] = df["Eleitores"].astype(int) - df["Eleitores_perfil"]

    # Local de residencia predominante e sua participacao.
    top = (residencia.sort_values("QT_ELEITORES", ascending=False)
           .groupby("NR_SECAO").first())
    df["Residencia_predominante"] = df["Secao"].map(top["NM_LOCAL_VOTACAO"])
    df["Eleitores_residencia_predominante"] = (
        df["Secao"].map(top["QT_ELEITORES"]).fillna(0).astype(int))

    return df.drop(columns=["Secao_principal_bruta"]).sort_values("Secao")


def monta_urnas(secoes: pd.DataFrame) -> pd.DataFrame:
    """Uma linha por urna: principal + agregada e o total combinado."""
    principais = secoes[secoes["Papel"] == "Principal"].set_index("Secao")
    agregadas = secoes[secoes["Papel"] == "Agregada"].set_index("Secao")

    linhas = []
    for urna, principal in principais.iterrows():
        filhas = agregadas[agregadas["Urna"] == urna]
        agregada = int(filhas.index[0]) if len(filhas) else None
        eleitores_agregada = int(filhas["Eleitores"].sum()) if len(filhas) else 0
        linhas.append({
            "Urna": int(urna),
            "Secao_principal": int(urna),
            "Secao_agregada": agregada,
            "Eleitores_principal": int(principal["Eleitores"]),
            "Eleitores_agregada": eleitores_agregada,
            "Total_combinado": int(principal["Eleitores"]) + eleitores_agregada,
            "Qtd_secoes": 1 + len(filhas),
        })

    df = pd.DataFrame(linhas).sort_values("Total_combinado", ascending=False)
    df.insert(0, "Posicao", range(1, len(df) + 1))
    return df.reset_index(drop=True)


def matriz_residencia(residencia: pd.DataFrame, secoes: pd.DataFrame,
                      por: str) -> pd.DataFrame:
    """Matriz eleitores x local de residencia, por secao ou por urna."""
    base = residencia.merge(
        secoes[["Secao", "Urna"]], left_on="NR_SECAO", right_on="Secao")
    chave = "NR_SECAO" if por == "secao" else "Urna"

    mat = base.pivot_table(index=chave, columns="NM_LOCAL_VOTACAO",
                           values="QT_ELEITORES", aggfunc="sum",
                           fill_value=0).astype(int)

    # DUBLIN primeiro, o resto por volume decrescente.
    ordem = mat.sum().sort_values(ascending=False).index.tolist()
    if "DUBLIN" in ordem:
        ordem.insert(0, ordem.pop(ordem.index("DUBLIN")))
    mat = mat[ordem]

    mat["TOTAL"] = mat.sum(axis=1)
    mat = mat.sort_values("TOTAL", ascending=False)
    mat.index.name = "Secao" if por == "secao" else "Urna"
    mat.loc["TOTAL"] = mat.sum()
    return mat.reset_index()


def monta_inconsistencias(locais: pd.DataFrame, secoes: pd.DataFrame,
                          urnas: pd.DataFrame,
                          divergencias_mapa: list) -> pd.DataFrame:
    """Registra tudo o que nao fechou, em vez de esconder."""
    itens = []

    for d in divergencias_mapa:
        itens.append({"Tipo": "Divergencia PNG x CSV", "Referencia": d["ref"],
                      "Descricao": d["desc"]})

    difs = secoes[secoes["Diferenca"] != 0]
    if len(difs):
        for _, r in difs.iterrows():
            itens.append({
                "Tipo": "Divergência entre as duas fontes",
                "Referencia": f"Seção {int(r['Secao'])}",
                "Descricao": (
                    f"QT_ELEITOR_SECAO={int(r['Eleitores'])} x "
                    f"perfil={int(r['Eleitores_perfil'])} "
                    f"(diferença {int(r['Diferenca'])})"),
            })
    else:
        itens.append({
            "Tipo": "Reconciliação",
            "Referencia": "Todas as seções",
            "Descricao": ("As duas fontes batem seção a seção: "
                          f"{int(secoes['Eleitores'].sum()):,} eleitores.")
                         .replace(",", "."),
        })

    orfas = set(locais["NR_SECAO"].astype(int)) - set(secoes["Secao"].astype(int))
    for s in sorted(orfas):
        itens.append({"Tipo": "Seção órfã", "Referencia": f"Seção {s}",
                      "Descricao": "Não foi possível atribuir a uma urna."})

    sozinhas = urnas[urnas["Secao_agregada"].isna()]
    itens.append({
        "Tipo": "Nota",
        "Referencia": f"{len(sozinhas)} urnas sem seção agregada",
        "Descricao": ("As seções " + ", ".join(
            str(int(s)) for s in sorted(sozinhas["Secao_principal"]))
            + " operam sozinhas na urna."),
    })
    itens.append({
        "Tipo": "Nota",
        "Referencia": "Datas de geração",
        "Descricao": ("Locais de votação gerado em 13/08/2026; perfil do "
                      "eleitorado em 14/07/2026."),
    })
    return pd.DataFrame(itens)


def escreve_aba(writer, df: pd.DataFrame, nome: str, escala_col: str = None):
    """Grava uma aba ja formatada."""
    df.to_excel(writer, sheet_name=nome, index=False)
    ws = writer.sheets[nome]

    for celula in ws[1]:
        celula.fill = CABECALHO
        celula.font = FONTE_CABECALHO
        celula.alignment = Alignment(horizontal="center", vertical="center",
                                     wrap_text=True)
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    for i, coluna in enumerate(df.columns, start=1):
        largura = max(len(str(coluna)),
                      *(len(str(v)) for v in df[coluna].head(200))) + 3
        ws.column_dimensions[get_column_letter(i)].width = min(largura, 34)

    if escala_col and escala_col in df.columns:
        letra = get_column_letter(list(df.columns).index(escala_col) + 1)
        faixa = f"{letra}2:{letra}{len(df) + 1}"
        ws.conditional_formatting.add(faixa, ColorScaleRule(
            start_type="min", start_color="F8696B",
            mid_type="percentile", mid_value=50, mid_color="FFEB84",
            end_type="max", end_color="63BE7B"))
    return ws


def main():
    SAIDAS.mkdir(exist_ok=True)

    locais = carrega_locais_votacao()
    perfil = carrega_perfil_eleitorado()
    residencia = residencia_por_secao(perfil)

    # Confere o mapa do PNG contra a coluna oficial NR_SECAO_PRINCIPAL.
    mapa_csv = {
        int(r["NR_SECAO"]): int(r["NR_SECAO_PRINCIPAL"])
        for _, r in locais[locais["DS_TIPO_SECAO_AGREGADA"] == "Agregada"].iterrows()
    }
    divergencias = []
    for agregada in sorted(set(MAPA_PNG) | set(mapa_csv)):
        no_png, no_csv = MAPA_PNG.get(agregada), mapa_csv.get(agregada)
        if no_png != no_csv:
            divergencias.append({
                "ref": f"Seção agregada {agregada}",
                "desc": (f"O PNG do TSE indica a seção principal {no_png}; o CSV "
                         f"oficial indica {no_csv}. A seção {no_png} não existe "
                         f"em Dublin — pertence ao PORTO —, então o PNG traz erro "
                         f"de digitação e o CSV prevalece."),
            })

    secoes = monta_secoes(locais, residencia)
    urnas = monta_urnas(secoes)
    mat_secao = matriz_residencia(residencia, secoes, "secao")
    mat_urna = matriz_residencia(residencia, secoes, "urna")
    inconsistencias = monta_inconsistencias(locais, secoes, urnas, divergencias)

    # --- Validacoes que precisam passar antes de gravar -------------------
    assert secoes["Eleitores"].sum() == urnas["Total_combinado"].sum(), \
        "Soma por secao difere da soma por urna"
    assert secoes["Eleitores"].sum() == perfil["QT_ELEITORES"].sum(), \
        "Soma das secoes difere do total do perfil"
    assert len(secoes) == len(locais), "Alguma secao se perdeu no caminho"
    assert secoes["Urna"].nunique() == len(urnas), "Contagem de urnas inconsistente"

    # Conferencia independente: o total que calculamos por urna tem de bater
    # com o campo que o proprio TSE ja traz na secao principal.
    tse = secoes.set_index("Secao")["Total_urna_TSE"]
    conferidas = sum(
        int(tse[r["Secao_principal"]]) == int(r["Total_combinado"])
        for _, r in urnas.iterrows()
    )
    assert conferidas == len(urnas), (
        f"Apenas {conferidas}/{len(urnas)} urnas batem com QT_ELEITOR_ELEICAO_FEDERAL")
    print(f"Conferencia TSE: {conferidas}/{len(urnas)} urnas batem com "
          f"QT_ELEITOR_ELEICAO_FEDERAL")

    caminho = SAIDAS / "Dublin_2026_agregacoes.xlsx"
    with pd.ExcelWriter(caminho, engine="openpyxl") as writer:
        escreve_aba(writer, urnas, "Urnas", escala_col="Total_combinado")
        escreve_aba(writer, secoes, "Secoes", escala_col="Eleitores")
        escreve_aba(writer, mat_secao, "Residencia x Secao")
        escreve_aba(writer, mat_urna, "Residencia x Urna")
        escreve_aba(writer, inconsistencias, "Inconsistencias")

    dados = {
        "local": "Royal Dublin Society - Hall 2",
        "endereco": "RDS, Merrion Road, Ballsbridge, Dublin 4 D04 AK83",
        "total_eleitores": int(secoes["Eleitores"].sum()),
        "total_secoes": int(len(secoes)),
        "total_urnas": int(len(urnas)),
        "urnas": urnas.where(pd.notna(urnas), None).to_dict("records"),
        "secoes": secoes.to_dict("records"),
        "residencia_urna": mat_urna[mat_urna["Urna"] != "TOTAL"].to_dict("records"),
        "residencia_total": (residencia.groupby("NM_LOCAL_VOTACAO")["QT_ELEITORES"]
                             .sum().sort_values(ascending=False).to_dict()),
        "inconsistencias": inconsistencias.to_dict("records"),
    }
    (SAIDAS / "dados.json").write_text(
        json.dumps(dados, ensure_ascii=False, indent=2, default=str),
        encoding="utf-8")

    print(f"Planilha : {caminho}")
    print(f"JSON     : {SAIDAS / 'dados.json'}")
    print(f"\n{len(urnas)} urnas | {len(secoes)} secoes | "
          f"{secoes['Eleitores'].sum():,} eleitores")
    print(f"\nTop 10 urnas por total combinado:")
    print(urnas.head(10).to_string(index=False))
    print(f"\nMenores 5:")
    print(urnas.tail(5).to_string(index=False))


if __name__ == "__main__":
    main()
